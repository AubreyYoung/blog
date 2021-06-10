import openpyxl

wb = openpyxl.load_workbook(filename='00.FusionSphere组网规划_inspur_20210428.xlsx')
print(wb.sheetnames)
sheet = wb['计算资源设计']
print(sheet)
print(sheet.dimensions)
cell1 = sheet["B2"]
cell2 = sheet["C3"]
print(cell1.value, cell2.value)

print('############################################################################')
cell3 = sheet.cell(row=1, column=1)
cell4 = sheet.cell(row=11, column=3)
print(cell3.value, cell4.value)

# cell = sheet["A1:E5"]
# print(cell)
# for i in cell:
#    for j in i:
#        print(j.value)
